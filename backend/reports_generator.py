import io
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def generate_pdf_report(report_type, data):
    """
    Generates a professional PDF report.
    report_type: 'daily', 'defaulter', 'leave', 'occupancy'
    data: list of dicts representing rows
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        leading=24,
        textColor=colors.HexColor('#1E3A8A'),
        alignment=1, # Center
        spaceAfter=10
    )
    
    subtitle_style = ParagraphStyle(
        'SubtitleStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#4B5563'),
        alignment=1,
        spaceAfter=20
    )
    
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=10,
        leading=12,
        textColor=colors.white
    )
    
    body_style = ParagraphStyle(
        'BodyStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=11,
        textColor=colors.HexColor('#1F2937')
    )

    # Add Title and Headers
    story.append(Paragraph("JIM HOSTEL ATTENDANCE MANAGEMENT SYSTEM", title_style))
    date_str = datetime.now().strftime("%d-%b-%Y %I:%M %p")
    story.append(Paragraph(f"Report Type: {report_type.upper()} REPORT | Generated on: {date_str}", subtitle_style))
    story.append(Spacer(1, 10))

    if not data:
        story.append(Paragraph("No records found for the selected criteria.", styles['Normal']))
    else:
        # Define table headers based on report type
        if report_type == 'daily':
            headers = ["Student Name", "Room", "Status", "Marked By", "Time"]
            col_widths = [2.4*inch, 0.8*inch, 1.0*inch, 2.0*inch, 1.8*inch]
            rows = [[Paragraph(h, header_style) for h in headers]]
            for item in data:
                rows.append([
                    Paragraph(item.get('name', ''), body_style),
                    Paragraph(item.get('room_number', ''), body_style),
                    Paragraph(item.get('status', ''), body_style),
                    Paragraph(item.get('marked_by', ''), body_style),
                    Paragraph(item.get('time', ''), body_style)
                ])
        elif report_type == 'defaulter':
            headers = ["Student Name", "Room", "Attendance %", "Risk Level"]
            col_widths = [3.0*inch, 1.2*inch, 1.8*inch, 2.0*inch]
            rows = [[Paragraph(h, header_style) for h in headers]]
            for item in data:
                rows.append([
                    Paragraph(item.get('name', ''), body_style),
                    Paragraph(item.get('room_number', ''), body_style),
                    Paragraph(f"{item.get('attendance_percentage', 100.0)}%", body_style),
                    Paragraph(item.get('risk_level', ''), body_style)
                ])
        elif report_type == 'leave':
            headers = ["Student Name", "Room", "Leave From", "Leave To", "Reason", "Status"]
            col_widths = [2.0*inch, 0.8*inch, 1.2*inch, 1.2*inch, 2.0*inch, 0.8*inch]
            rows = [[Paragraph(h, header_style) for h in headers]]
            for item in data:
                rows.append([
                    Paragraph(item.get('name', ''), body_style),
                    Paragraph(item.get('room_number', ''), body_style),
                    Paragraph(item.get('leave_from', ''), body_style),
                    Paragraph(item.get('leave_to', ''), body_style),
                    Paragraph(item.get('reason', ''), body_style),
                    Paragraph(item.get('status', ''), body_style)
                ])
        elif report_type == 'occupancy':
            headers = ["Room No", "Block", "Floor", "Capacity", "Occupied", "Available"]
            col_widths = [1.2*inch, 1.5*inch, 1.0*inch, 1.3*inch, 1.3*inch, 1.7*inch]
            rows = [[Paragraph(h, header_style) for h in headers]]
            for item in data:
                rows.append([
                    Paragraph(item.get('room_number', ''), body_style),
                    Paragraph(item.get('block', ''), body_style),
                    Paragraph(str(item.get('floor', '')), body_style),
                    Paragraph(str(item.get('capacity', '')), body_style),
                    Paragraph(str(item.get('occupied', '')), body_style),
                    Paragraph(str(item.get('available_beds', '')), body_style)
                ])
        else:
            headers = ["Info"]
            col_widths = [8.0*inch]
            rows = [[Paragraph("Invalid Report Type", body_style)]]

        # Draw Table
        table = Table(rows, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A8A')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ]))
        story.append(table)
        
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def generate_excel_report(report_type, data):
    """
    Generates a styled Excel spreadsheet.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{report_type.capitalize()} Report"
    
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    title_font = Font(name='Arial', size=16, bold=True, color='1E3A8A')
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
    body_font = Font(name='Arial', size=10)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    
    # Title Row
    ws.merge_cells('A1:F1')
    ws['A1'] = "JIM HOSTEL ATTENDANCE MANAGEMENT SYSTEM"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align
    ws.row_dimensions[1].height = 40
    
    # Subtitle Row
    ws.merge_cells('A2:F2')
    date_str = datetime.now().strftime("%d-%b-%Y %I:%M %p")
    ws['A2'] = f"Report Type: {report_type.upper()} | Generated: {date_str}"
    ws['A2'].font = Font(name='Arial', size=10, italic=True, color='4B5563')
    ws['A2'].alignment = center_align
    ws.row_dimensions[2].height = 20
    
    ws.append([]) # Spacer row

    if not data:
        ws.append(["No records found for the selected criteria."])
        ws.merge_cells('A4:F4')
    else:
        # Define headers
        if report_type == 'daily':
            headers = ["Student Name", "Room Number", "Attendance Status", "Marked By", "Timestamp"]
            ws.append(headers)
            for item in data:
                ws.append([
                    item.get('name', ''),
                    item.get('room_number', ''),
                    item.get('status', ''),
                    item.get('marked_by', ''),
                    item.get('time', '')
                ])
        elif report_type == 'defaulter':
            headers = ["Student Name", "Room Number", "Attendance Percentage", "Risk Level"]
            ws.append(headers)
            for item in data:
                ws.append([
                    item.get('name', ''),
                    item.get('room_number', ''),
                    f"{item.get('attendance_percentage', 100.0)}%",
                    item.get('risk_level', '')
                ])
        elif report_type == 'leave':
            headers = ["Student Name", "Room Number", "Leave From", "Leave To", "Reason", "Status"]
            ws.append(headers)
            for item in data:
                ws.append([
                    item.get('name', ''),
                    item.get('room_number', ''),
                    item.get('leave_from', ''),
                    item.get('leave_to', ''),
                    item.get('reason', ''),
                    item.get('status', '')
                ])
        elif report_type == 'occupancy':
            headers = ["Room Number", "Block", "Floor", "Capacity", "Occupied Beds", "Available Beds"]
            ws.append(headers)
            for item in data:
                ws.append([
                    item.get('room_number', ''),
                    item.get('block', ''),
                    item.get('floor', 0),
                    item.get('capacity', 0),
                    item.get('occupied', 0),
                    item.get('available_beds', 0)
                ])

        # Style the headers (Row 4 is the header row now since Row 1 is Title, Row 2 is Subtitle, Row 3 is Spacer)
        header_row_idx = 4
        ws.row_dimensions[header_row_idx].height = 25
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row_idx, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Style body rows
        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 20
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = body_font
                cell.border = thin_border
                cell.alignment = left_align
                
                # Alternate row shading
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            
            # Skip rows 1 and 2 when calculating lengths
            for cell in col[2:]:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
